# -*- coding: utf-8 -*-
from __future__ import annotations

import argparse
import os
import re
import unicodedata
from datetime import datetime
from pathlib import Path
from typing import Iterable

import pandas as pd
from openpyxl import load_workbook


APP_DIR = Path(__file__).resolve().parents[1]
ROOT_DIR = Path(__file__).resolve().parents[4]
DATA_DIR = APP_DIR / "data"
OUTPUT_EXCEL = DATA_DIR / "patrimonio_biodiversidad_base.xlsx"

ESTUDIOS_DIR = ROOT_DIR / "ESTUDIOS"
YEAR_CONFIG = {
    2024: {
        "base_dir": ESTUDIOS_DIR / "ESTUDIOS_2024",
        "master": ESTUDIOS_DIR / "ESTUDIOS_2024" / "Estudios de Investigación 2024.xlsx",
    },
    2025: {
        "base_dir": ESTUDIOS_DIR / "ESTUDIOS_2025",
        "master": ESTUDIOS_DIR / "ESTUDIOS_2025" / "Estudios_de_Investigación_2025.xlsx",
    },
}

EXCEL_EXTENSIONS = {".xlsx", ".xlsm", ".xls", ".csv"}
OPENPYXL_EXTENSIONS = {".xlsx", ".xlsm"}
ROOT_CONTROL_PREFIXES = ("LOG_",)

SPECIES_HINTS = {
    "especie",
    "nombre cientifico",
    "nombre científico",
    "taxon",
    "taxón",
    "familia",
    "orden",
}
LOCATION_HINTS = {
    "departamento",
    "provincia",
    "distrito",
    "localidad",
    "estacion",
    "estación",
    "unidad de vegetacion",
    "unidad de vegetación",
}
COORD_HINTS = {"este", "norte", "utm", "coordenada", "wgs"}
GROUP_HINTS = {
    "flora": "Flora",
    "vegetacion": "Flora",
    "vegetación": "Flora",
    "aves": "Fauna",
    "ave": "Fauna",
    "mamifer": "Fauna",
    "mastofauna": "Fauna",
    "reptil": "Fauna",
    "anfib": "Fauna",
    "herpeto": "Fauna",
    "artrop": "Fauna",
    "insect": "Fauna",
    "fauna": "Fauna",
}

PERU_DEPARTMENTS = {
    "amazonas": "Amazonas",
    "ancash": "Ancash",
    "apurimac": "Apurimac",
    "arequipa": "Arequipa",
    "ayacucho": "Ayacucho",
    "cajamarca": "Cajamarca",
    "cusco": "Cusco",
    "huancavelica": "Huancavelica",
    "huanuco": "Huanuco",
    "ica": "Ica",
    "junin": "Junin",
    "la libertad": "La Libertad",
    "lambayeque": "Lambayeque",
    "lima": "Lima",
    "loreto": "Loreto",
    "madre de dios": "Madre de Dios",
    "moquegua": "Moquegua",
    "pasco": "Pasco",
    "piura": "Piura",
    "puno": "Puno",
    "san martin": "San Martin",
    "tacna": "Tacna",
    "tumbes": "Tumbes",
    "ucayali": "Ucayali",
}


def normalize_text(value: object) -> str:
    if pd.isna(value):
        return ""
    text = str(value).strip()
    text = unicodedata.normalize("NFKD", text)
    text = "".join(ch for ch in text if not unicodedata.combining(ch))
    text = re.sub(r"\s+", " ", text)
    return text.strip()


def normalize_key(value: object) -> str:
    text = normalize_text(value).lower()
    text = re.sub(r"[^a-z0-9]+", "_", text)
    return text.strip("_")


def normalize_department(value: object) -> str:
    text = normalize_text(value)
    if not text:
        return ""
    key = normalize_text(text).lower()
    key = key.replace(" y ", ",")
    key = key.replace("/", ",")
    parts = [part.strip() for part in re.split(r"[,;]", key) if part.strip()]
    found: list[str] = []
    for part in parts:
        part_key = normalize_text(part).lower()
        if part_key in {"callao", "provincia constitucional del callao"}:
            continue
        for dep_key, dep_name in PERU_DEPARTMENTS.items():
            if part_key == dep_key or dep_key in part_key:
                if dep_name not in found:
                    found.append(dep_name)
    return "; ".join(found)


def first_existing(row: pd.Series, candidates: Iterable[str]) -> str:
    for candidate in candidates:
        if candidate in row.index:
            value = normalize_text(row.get(candidate))
            if value:
                return value
    return ""


def stable_code(value: object, year: int, index: int) -> str:
    text = normalize_text(value)
    match = re.search(r"(\d{1,3})\s*$", text)
    if match:
        return f"{year}-{int(match.group(1)):03d}"
    match = re.search(r"(\d{1,3})", text)
    if match:
        return f"{year}-{int(match.group(1)):03d}"
    return f"{year}-{index:03d}"


def infer_study_code_from_path(path: Path, base_dir: Path, year: int) -> str:
    try:
        rel = path.relative_to(base_dir)
        top = rel.parts[0]
    except ValueError:
        return ""
    match = re.search(r"(?:M_)?EST\d{2}[_\-\s]*(\d{3})", top, flags=re.IGNORECASE)
    if match:
        return f"{year}-{match.group(1)}"
    match = re.match(r"(\d{1,3})", top)
    if match:
        return f"{year}-{int(match.group(1)):03d}"
    return top


def normalize_master_columns(df: pd.DataFrame) -> pd.DataFrame:
    columns: list[str] = []
    blank_count = 0
    for col in df.columns:
        name = normalize_text(col)
        if not name or name.lower().startswith("unnamed"):
            blank_count += 1
            name = "instrumento_fuente" if blank_count == 1 else f"campo_sin_nombre_{blank_count}"
        columns.append(name)
    df = df.copy()
    df.columns = columns
    return df


def load_master_sources(year: int, master_path: Path) -> pd.DataFrame:
    if not master_path.exists():
        return pd.DataFrame()
    df = pd.read_excel(master_path, sheet_name=0, dtype=object, engine="openpyxl")
    df = normalize_master_columns(df).dropna(how="all")
    rows = []
    for idx, row in df.iterrows():
        title = first_existing(row, ["Titulo", "Título", "titulo", "título"])
        if not title:
            continue
        numeracion = first_existing(row, ["Numeración", "Numeracion", "numero_registro"])
        code = stable_code(numeracion, year, len(rows) + 1)
        departamento_original = first_existing(row, ["Departamento", "departamento"])
        rows.append(
            {
                "id_fuente": code,
                "anio": year,
                "numeracion": numeracion,
                "nro_expediente": first_existing(row, ["Nro Expediente", "N° Expediente", "Expediente"]),
                "titulo_fuente": title,
                "tipo_documento_original": first_existing(row, ["Tipo de documento", "Tipo Documento"]),
                "tipo_documento_normalizado": normalize_document_type(
                    first_existing(row, ["Tipo de documento", "Tipo Documento"])
                ),
                "instrumento_fuente": normalize_instrument(first_existing(row, ["instrumento_fuente", "Tipo", "tipo"])),
                "autor_institucion": first_existing(row, ["Remitente", "Autor", "Institución", "Institucion"]),
                "departamento": departamento_original,
                "departamento_normalizado": normalize_department(departamento_original),
                "provincia": first_existing(row, ["Provincia", "provincia"]),
                "resumen_fuente": first_existing(row, ["Resumen", "resumen"]),
                "archivo_maestro": str(master_path),
                "fila_maestro": int(idx) + 2,
                "estado_revision": "pendiente_revision" if not first_existing(row, ["Departamento"]) else "inicial_excel",
            }
        )
    return pd.DataFrame(rows)


def normalize_document_type(value: object) -> str:
    text = normalize_text(value).lower()
    if not text:
        return ""
    if "cumpl" in text and "final" in text:
        return "Informe final de cumplimiento"
    if "cumpl" in text:
        return "Informe de cumplimiento"
    if "parcial" in text:
        return "Informe parcial"
    if "final" in text:
        return "Informe final"
    return normalize_text(value)


def normalize_instrument(value: object) -> str:
    text = normalize_text(value)
    key = normalize_key(text)
    if key == "iga":
        return "Instrumentos de Gestión Ambiental"
    if key in {"autorizacion_de_investigacion", "autorizacion_investigacion"}:
        return "Autorización de Investigación"
    return text


def classify_group(text: str) -> tuple[str, str]:
    raw = text.lower()
    for hint, group in GROUP_HINTS.items():
        if hint in raw:
            subgroup = ""
            if "ave" in raw:
                subgroup = "Ave"
            elif "mamifer" in raw or "masto" in raw:
                subgroup = "Mamífero"
            elif "reptil" in raw or "anfib" in raw or "herpeto" in raw:
                subgroup = "Herpetofauna"
            elif "artrop" in raw or "insect" in raw:
                subgroup = "Artrópodo"
            elif "flora" in raw or "veget" in raw:
                subgroup = "Flora"
            return group, subgroup
    return "", ""


def list_tabular_files(base_dir: Path) -> list[Path]:
    if not base_dir.exists():
        return []
    files: list[Path] = []
    for root, _, filenames in os.walk(base_dir, onerror=lambda _: None):
        root_path = Path(root)
        for filename in filenames:
            path = root_path / filename
            if path.name.startswith("~$"):
                continue
            if path.suffix.lower() not in EXCEL_EXTENSIONS:
                continue
            if path.parent == base_dir and path.name.upper().startswith(ROOT_CONTROL_PREFIXES):
                continue
            files.append(path)
    return sorted(files)


def inspect_workbook(path: Path) -> list[dict[str, object]]:
    ext = path.suffix.lower()
    if ext == ".csv":
        return [{"hoja": "CSV", "filas": None, "columnas": None, "encabezado_detectado": "", "estado_lectura": "csv"}]
    if ext not in OPENPYXL_EXTENSIONS:
        return [{"hoja": "", "filas": None, "columnas": None, "encabezado_detectado": "", "estado_lectura": "formato_no_openpyxl"}]
    try:
        wb = load_workbook(path, read_only=True, data_only=True)
    except Exception as exc:  # noqa: BLE001
        return [{"hoja": "", "filas": None, "columnas": None, "encabezado_detectado": "", "estado_lectura": f"error: {exc}"}]

    records: list[dict[str, object]] = []
    try:
        for ws in wb.worksheets:
            rows = []
            for row in ws.iter_rows(min_row=1, max_row=8, values_only=True):
                rows.append([normalize_text(cell) for cell in row])
            header = guess_header(rows)
            records.append(
                {
                    "hoja": ws.title,
                    "filas": ws.max_row,
                    "columnas": ws.max_column,
                    "encabezado_detectado": " | ".join([h for h in header if h][:40]),
                    "estado_lectura": "ok",
                }
            )
    finally:
        wb.close()
    return records


def guess_header(rows: list[list[str]]) -> list[str]:
    best: list[str] = []
    best_score = -1
    for row in rows:
        compact = [cell for cell in row if cell]
        joined = " ".join(compact).lower()
        score = len(compact)
        if any(h in joined for h in SPECIES_HINTS | LOCATION_HINTS | COORD_HINTS):
            score += 20
        if score > best_score:
            best = row
            best_score = score
    return best


def guess_header_with_index(rows: list[list[str]]) -> tuple[int, list[str]]:
    best: list[str] = []
    best_index = 0
    best_score = -1
    for idx, row in enumerate(rows, start=1):
        compact = [cell for cell in row if cell]
        joined = " ".join(compact).lower()
        score = len(compact)
        if any(h in joined for h in SPECIES_HINTS | LOCATION_HINTS | COORD_HINTS):
            score += 20
        if score > best_score:
            best = row
            best_index = idx
            best_score = score
    return best_index, best


def find_column(header: list[str], candidates: Iterable[str]) -> int | None:
    normalized = [normalize_key(cell) for cell in header]
    for candidate in candidates:
        key = normalize_key(candidate)
        for idx, col in enumerate(normalized):
            if col == key or key in col:
                return idx
    return None


def value_at(row: tuple[object, ...], idx: int | None) -> str:
    if idx is None or idx >= len(row):
        return ""
    return normalize_text(row[idx])


def extract_species_records(
    path: Path,
    sheet_name: str,
    *,
    year: int,
    id_fuente: str,
    file_id: int,
    sheet_idx: int,
    group_default: str,
    subgroup_default: str,
) -> list[dict[str, object]]:
    if path.suffix.lower() not in OPENPYXL_EXTENSIONS:
        return []
    try:
        wb = load_workbook(path, read_only=True, data_only=True)
    except Exception:
        return []
    records: list[dict[str, object]] = []
    try:
        ws = wb[sheet_name]
        preview = [[normalize_text(cell) for cell in row] for row in ws.iter_rows(min_row=1, max_row=8, values_only=True)]
        header_row, header = guess_header_with_index(preview)
        species_idx = find_column(
            header,
            [
                "nombre cientifico",
                "nombre científico",
                "especie nombre cientifico",
                "especie nombre científico",
                "nombre cientifico ",
                "especie",
                "especies",
                "taxon",
                "taxón",
            ],
        )
        if species_idx is None:
            return []

        common_idx = find_column(header, ["nombre comun", "nombre común", "nombre local"])
        family_idx = find_column(header, ["familia"])
        order_idx = find_column(header, ["orden"])
        class_idx = find_column(header, ["clase"])
        department_idx = find_column(header, ["departamento", "region", "región"])
        province_idx = find_column(header, ["provincia"])
        district_idx = find_column(header, ["distrito"])
        station_idx = find_column(header, ["estacion", "estación", "estacion de muestreo", "estación de muestreo"])
        unit_idx = find_column(header, ["unidad de vegetacion", "unidad de vegetación", "cobertura vegetal"])
        east_idx = find_column(header, ["este", "este utm", "este i", "este m"])
        north_idx = find_column(header, ["norte", "norte utm", "norte i", "norte m"])
        utm_idx = find_column(header, ["utm", "zona utm", "coordenadas utm wgs84", "wgs 84"])
        record_type_idx = find_column(header, ["tipo registro", "tipo de registro", "tipo. reg.", "registro"])
        count_idx = find_column(header, ["numero de individuos", "número de individuos", "individuos", "abundancia"])
        season_idx = find_column(header, ["temporada", "epoca", "época"])
        method_idx = find_column(header, ["metodo", "método", "metodologia", "metodología"])
        group_idx = find_column(header, ["grupo taxonomico", "grupo taxonómico", "grupo biologico", "grupo biológico"])

        for excel_row, row in enumerate(ws.iter_rows(min_row=header_row + 1, values_only=True), start=header_row + 1):
            scientific_name = value_at(row, species_idx)
            if not scientific_name or scientific_name.lower() in {"especie", "especies", "total", "total general"}:
                continue
            if re.fullmatch(r"\d+(\.\d+)?", scientific_name):
                continue
            row_text = " ".join(normalize_text(cell) for cell in row if normalize_text(cell))
            row_group, row_subgroup = classify_group(f"{group_default} {subgroup_default} {value_at(row, group_idx)} {sheet_name}")
            records.append(
                {
                    "id_registro": f"{year}-XREG-{file_id:05d}-{sheet_idx:03d}-{excel_row:06d}",
                    "id_fuente": id_fuente,
                    "anio": year,
                    "origen_dato": "excel",
                    "nivel_registro": "fila_extraida",
                    "archivo_origen": str(path),
                    "hoja_origen": sheet_name,
                    "fila_origen": excel_row,
                    "grupo_general": row_group or group_default,
                    "subgrupo": row_subgroup or subgroup_default,
                    "ambito_ecologico": "",
                    "clase": value_at(row, class_idx),
                    "orden": value_at(row, order_idx),
                    "familia": value_at(row, family_idx),
                    "nombre_cientifico": scientific_name,
                    "nombre_comun": value_at(row, common_idx),
                    "departamento": value_at(row, department_idx),
                    "provincia": value_at(row, province_idx),
                    "distrito": value_at(row, district_idx),
                    "unidad_ecosistemica": value_at(row, unit_idx),
                    "estacion": value_at(row, station_idx),
                    "este": value_at(row, east_idx),
                    "norte": value_at(row, north_idx),
                    "zona_utm": value_at(row, utm_idx),
                    "temporada": value_at(row, season_idx),
                    "metodo_registro": value_at(row, method_idx),
                    "tipo_registro": value_at(row, record_type_idx) or "reporte",
                    "numero_individuos_original": value_at(row, count_idx),
                    "conteo_reportes": 1,
                    "estado_revision": "extraido_excel_pendiente_validacion",
                    "observaciones": "Extracción heurística desde Excel; validar columnas y especie.",
                    "texto_fila_origen": row_text[:500],
                }
            )
    finally:
        wb.close()
    return records


def is_biodiversity_like(text: str) -> bool:
    key = text.lower()
    hints = SPECIES_HINTS | LOCATION_HINTS | COORD_HINTS | set(GROUP_HINTS)
    return any(h in key for h in hints)


def build_inventory(year: int, base_dir: Path) -> tuple[pd.DataFrame, pd.DataFrame, pd.DataFrame]:
    file_rows = []
    sheet_rows = []
    candidate_rows = []
    for file_id, path in enumerate(list_tabular_files(base_dir), start=1):
        rel = str(path.relative_to(base_dir))
        study_folder = rel.split("\\")[0]
        id_fuente = infer_study_code_from_path(path, base_dir, year)
        file_text = f"{path.name} {study_folder}"
        group, subgroup = classify_group(file_text)
        file_rows.append(
            {
                "id_archivo_excel": f"{year}-XLS-{file_id:05d}",
                "anio": year,
                "id_fuente": id_fuente,
                "carpeta_estudio": study_folder,
                "archivo_excel": path.name,
                "ruta_archivo": str(path),
                "extension": path.suffix.lower(),
                "tamano_bytes": path.stat().st_size if path.exists() else None,
                "fecha_modificacion": datetime.fromtimestamp(path.stat().st_mtime).isoformat(timespec="seconds")
                if path.exists()
                else "",
                "grupo_sugerido_archivo": group,
                "subgrupo_sugerido_archivo": subgroup,
            }
        )
        for sheet_idx, sheet in enumerate(inspect_workbook(path), start=1):
            sheet_text = f"{file_text} {sheet.get('hoja', '')} {sheet.get('encabezado_detectado', '')}"
            sheet_group, sheet_subgroup = classify_group(sheet_text)
            bio = is_biodiversity_like(sheet_text)
            sheet_id = f"{year}-SHT-{file_id:05d}-{sheet_idx:03d}"
            sheet_rows.append(
                {
                    "id_hoja_excel": sheet_id,
                    "id_archivo_excel": f"{year}-XLS-{file_id:05d}",
                    "anio": year,
                    "id_fuente": id_fuente,
                    "archivo_excel": path.name,
                    "hoja_excel": sheet.get("hoja", ""),
                    "filas_detectadas": sheet.get("filas"),
                    "columnas_detectadas": sheet.get("columnas"),
                    "encabezado_detectado": sheet.get("encabezado_detectado", ""),
                    "es_candidata_biodiversidad": "si" if bio else "no",
                    "grupo_sugerido": sheet_group or group,
                    "subgrupo_sugerido": sheet_subgroup or subgroup,
                    "estado_lectura": sheet.get("estado_lectura", ""),
                }
            )
            if bio:
                extracted = extract_species_records(
                    path,
                    str(sheet.get("hoja", "")),
                    year=year,
                    id_fuente=id_fuente,
                    file_id=file_id,
                    sheet_idx=sheet_idx,
                    group_default=sheet_group or group,
                    subgroup_default=sheet_subgroup or subgroup,
                )
                if extracted:
                    candidate_rows.extend(extracted)
                else:
                    candidate_rows.append(
                        {
                            "id_registro": f"{year}-CAND-{len(candidate_rows) + 1:06d}",
                            "id_fuente": id_fuente,
                            "anio": year,
                            "origen_dato": "excel",
                            "nivel_registro": "hoja_candidata",
                            "archivo_origen": str(path),
                            "hoja_origen": sheet.get("hoja", ""),
                            "fila_origen": "",
                            "grupo_general": sheet_group or group,
                            "subgrupo": sheet_subgroup or subgroup,
                            "ambito_ecologico": "",
                            "clase": "",
                            "orden": "",
                            "familia": "",
                            "nombre_cientifico": "",
                            "nombre_comun": "",
                            "departamento": "",
                            "provincia": "",
                            "distrito": "",
                            "unidad_ecosistemica": "",
                            "estacion": "",
                            "este": "",
                            "norte": "",
                            "zona_utm": "",
                            "temporada": "",
                            "metodo_registro": "",
                            "tipo_registro": "pendiente_extraccion_excel",
                            "numero_individuos_original": "",
                            "conteo_reportes": 0,
                            "estado_revision": "pendiente_extraccion",
                            "observaciones": "Hoja candidata detectada por encabezados/nombre de archivo; falta extraer filas.",
                            "texto_fila_origen": "",
                        }
                    )
    return pd.DataFrame(file_rows), pd.DataFrame(sheet_rows), pd.DataFrame(candidate_rows)


def merge_sources(sources: pd.DataFrame, inventory: pd.DataFrame) -> pd.DataFrame:
    if sources.empty:
        return sources
    counts = (
        inventory.groupby("id_fuente", dropna=False)
        .agg(archivos_excel=("id_archivo_excel", "nunique"))
        .reset_index()
        if not inventory.empty
        else pd.DataFrame(columns=["id_fuente", "archivos_excel"])
    )
    out = sources.merge(counts, on="id_fuente", how="left")
    out["archivos_excel"] = out["archivos_excel"].fillna(0).astype(int)
    return out


def build_quality(sources: pd.DataFrame, inventory: pd.DataFrame, sheets: pd.DataFrame, candidates: pd.DataFrame) -> pd.DataFrame:
    return pd.DataFrame(
        [
            {"control": "fuentes_maestras", "estado": "OK" if not sources.empty else "REVISAR", "n": len(sources)},
            {"control": "archivos_excel_inventariados", "estado": "OK" if not inventory.empty else "REVISAR", "n": len(inventory)},
            {"control": "hojas_excel_inventariadas", "estado": "OK" if not sheets.empty else "REVISAR", "n": len(sheets)},
            {
                "control": "hojas_candidatas_biodiversidad",
                "estado": "OK" if not candidates.empty else "REVISAR",
                "n": len(candidates),
            },
            {
                "control": "fuentes_sin_departamento",
                "estado": "REVISAR" if (not sources.empty and sources["departamento"].eq("").any()) else "OK",
                "n": int(sources["departamento"].eq("").sum()) if not sources.empty else 0,
            },
        ]
    )


def field_dictionary() -> pd.DataFrame:
    rows = [
        ("fuentes", "id_fuente", "Identificador anual de fuente/estudio."),
        ("fuentes", "instrumento_fuente", "Categoría institucional rescatada, por ejemplo Instrumentos de Gestión Ambiental o Autorización de Investigación."),
        ("inventario_excel", "ruta_archivo", "Ruta absoluta del Excel original inventariado."),
        ("hojas_excel", "encabezado_detectado", "Fila probable de encabezados detectada sin alterar el archivo original."),
        ("registros_especies", "nivel_registro", "Indica si el registro es una hoja candidata o una fila extraída."),
        ("registros_especies", "origen_dato", "Excel ahora; documento/PDF en avances posteriores."),
        ("control_calidad", "estado", "OK o REVISAR según control aplicado."),
    ]
    return pd.DataFrame(rows, columns=["tabla", "campo", "descripcion"])


def write_output(
    output: Path,
    sources: pd.DataFrame,
    inventory: pd.DataFrame,
    sheets: pd.DataFrame,
    candidates: pd.DataFrame,
    quality: pd.DataFrame,
) -> None:
    output.parent.mkdir(parents=True, exist_ok=True)
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        sources.to_excel(writer, sheet_name="01_fuentes", index=False)
        inventory.to_excel(writer, sheet_name="02_inventario_excel", index=False)
        sheets.to_excel(writer, sheet_name="03_hojas_excel", index=False)
        candidates.to_excel(writer, sheet_name="04_registros_especies", index=False)
        quality.to_excel(writer, sheet_name="05_control_calidad", index=False)
        field_dictionary().to_excel(writer, sheet_name="06_diccionario", index=False)


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Construye base trazable inicial desde Excel 2024-2025.")
    parser.add_argument("--years", default="2024,2025", help="Años separados por coma.")
    parser.add_argument("--output", default=str(OUTPUT_EXCEL), help="Ruta del Excel de salida.")
    return parser.parse_args()


def main() -> None:
    args = parse_args()
    years = [int(year.strip()) for year in args.years.split(",") if year.strip()]
    source_frames = []
    inventory_frames = []
    sheet_frames = []
    candidate_frames = []

    for year in years:
        config = YEAR_CONFIG[year]
        source_frames.append(load_master_sources(year, config["master"]))
        inventory, sheets, candidates = build_inventory(year, config["base_dir"])
        inventory_frames.append(inventory)
        sheet_frames.append(sheets)
        candidate_frames.append(candidates)

    sources = pd.concat(source_frames, ignore_index=True) if source_frames else pd.DataFrame()
    inventory = pd.concat(inventory_frames, ignore_index=True) if inventory_frames else pd.DataFrame()
    sheets = pd.concat(sheet_frames, ignore_index=True) if sheet_frames else pd.DataFrame()
    candidates = pd.concat(candidate_frames, ignore_index=True) if candidate_frames else pd.DataFrame()
    sources = merge_sources(sources, inventory)
    quality = build_quality(sources, inventory, sheets, candidates)
    write_output(Path(args.output), sources, inventory, sheets, candidates, quality)

    print(f"Base generada: {Path(args.output)}")
    print(f"Fuentes: {len(sources):,}")
    print(f"Archivos Excel/CSV: {len(inventory):,}")
    print(f"Hojas inventariadas: {len(sheets):,}")
    print(f"Hojas candidatas: {len(candidates):,}")


if __name__ == "__main__":
    main()
